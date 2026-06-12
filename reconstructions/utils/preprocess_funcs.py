import numpy as np
import pandas as pd
import pickle
import os
from sklearn.feature_selection import VarianceThreshold
from warnings import simplefilter
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from collections import defaultdict
from vedo import Tube, Sphere

MIDLINEZ = 5750
MIDLINEZ_10UM = 570

def rem_zero_var(df):
    '''
    removes columns with 0 variance from a df, might not be necessary if my json reading is working correctly
    
    :param df: DataFrame to have colums dropped from
    '''
    selector = VarianceThreshold(threshold=0.0).set_output(transform='pandas')
    df = selector.fit_transform(df)
    return df


def preprocess(df, log1p=True, pct=True):
    '''
    preprocess data however I want to, accepts a pandas.DataFrame with no NaNs
    this is written to have cells as cols and regions as rows i think

    :param df: pandas.DataFrame

    :param log1p: preprocessing option, default True, if False skips log1p scaling

    :param pct: preprocessing option, default True, if False skips cell size scaling

    returns: pandas.DataFrame
    '''
    data_tonorm = df.copy()

    #natural log scale my data, formula of ln(number of terminals + 1) as per Ding et al. 2025
    #and control for cell size
    #default behavior
    if log1p and pct:
        df_log = np.log(data_tonorm+1)
        col_sums = df_log.sum(axis=0)
        df_pct = (df_log.div(col_sums, axis=1))*100
        return df_pct
    
    if log1p and not pct:
        df_log = np.log(data_tonorm+1)
        return df_log
    
    if not log1p  and pct:
        row_sums = df.sum(axis=1)
        df_pct = (df.div(row_sums, axis=0))*100
        return df_pct
    
    if not log1p and not pct:
        #print('what are you using this for')
        return df
    
def merge_regions(df):
    '''
    merge ipsilateral/contralateral regions into one column
    (GPT assist)

    :param df: DataFrame with frequency information for cells, columns must be 'Ipsilateral [region]' or 'Contralateral [region]'

    returns: DataFrame with shape (n cells, n regions)
    '''
    #this suppresses the PerformanceWarning that pd throws, this function is still running fast
    simplefilter(action='ignore', category=pd.errors.PerformanceWarning)
    
    ipsi_cols = pd.Series(col for col in df.columns if col.startswith('Ipsilateral'))
    contra_cols = pd.Series(col for col in df.columns if col.startswith('Contralateral'))
    ipsi_regions = ipsi_cols.str.replace('Ipsilateral ', '')
    contra_regions = contra_cols.str.replace('Contrlateral ', '')
    all_regions = set(ipsi_regions) | set(contra_regions)
    
    merged_frequency = pd.DataFrame(index=df.index)
    for region in all_regions:
        ipsi_col = f'Ipsilateral {region}'
        contra_col = f'Contralateral {region}'
        ipsi_series = df[ipsi_col] if ipsi_col in df.columns else None
        contra_series = df[contra_col] if contra_col in df.columns else None
        
        #regions that recieve both ipsi and contra projections
        if ipsi_series is not None and contra_series is not None:
            merged_frequency[region] = df[ipsi_col] + df[contra_col]
            
        #regions that recieve only ipsilateral
        if ipsi_series is not None and contra_series is None:
            merged_frequency[region] = df[ipsi_col]
            
        #regions that only recieve contra
        if ipsi_series is None and contra_series is not None:
            merged_frequency[region] = df[contra_col]
    
    return merged_frequency

def get_df_for_region(df, region):
    '''
    get a df containing lateralized frequency data for a given target region

    :params df: full dataframe with columns as regions, lateralized, and rows as neurons

    :params region: target region as str
    '''
    ipsireg = 'Ipsilateral '+region
    contrareg = 'Contralateral '+region
    fullreg = [ipsireg, contrareg]

    regdf = df[fullreg]
    for cell, val in regdf.iterrows():
        iv, cv = val
        if iv == 0 and cv == 0:
            regdf = regdf.drop(cell)

    return regdf

def get_nodes_in_region(cells, regions, parcellated=True, ontlevel='structure', kind=None, infunc=False):
    '''
    Docstring for get_nodes_in_region
    
    :param cells: Description
    :param regions: Description
    returns a list of nodeIDs that I can then use to pull the specific nodes from coordswapped that have coords for visualizaiton
    or if inputting the non parcellated jsons, will regurn a list of nodes in desired region
    '''
# =============================================================================
#     if infunc:
#         regions=regions[0]
# =============================================================================
        
    match kind:
        case 'bulk':
            nodes = []
            for _, axon in cells.items():
                for node in axon:
                    if parcellated:
                        try: 
                            if node[ontlevel] in regions:
                                nodes.append(node['sampleNumber'])
                        except KeyError:
                            print(node)
                    #if you want not parcellated, then regions has to be the ont id (numbers), if using parcellated can find the region abv
                    if not parcellated:
                        if node['allenId'] in regions:
                            nodes.append(node)
            return nodes
        case 'by_cell':
            cellstonodes = {}
            for cell, axon in cells.items():
                nodes = []
                for node in axon:
                    if parcellated:
                        try:
                            if node[ontlevel] in regions:
                                nodes.append(node['sampleNumber'])
                        except KeyError:
                            print(node)
                    if not parcellated:
                        if node['allenId'] in regions:
                            nodes.append(node)
                cellstonodes[cell] = nodes
            return cellstonodes

def tenmicron_to_one(nodedict, onemicronpath):
    '''
    think what i want to do here is read in the dictionary of cells:nodeIds that registered in a given area (will have pulled that info earlier)
    and return a nested dictionary of cells:nodes with the coordinates before transforming to 10 micron resolution
    '''
    coordswappeddict = {}
    for cell, nodes in nodedict.items():
        cellfile = cell+'.json'
        cellpath = os.path.join(onemicronpath,cellfile)
        onemicron_nodes = []
        if nodes:
            with open(cellpath, 'r') as coordswappedcell:
                coordswapped = json.load(coordswappedcell)
                swappednodes = coordswapped['neurons'][0]['axon']
                for node in swappednodes:
                    if node['sampleNumber'] in nodes:
                        onemicron_nodes.append(node)
        coordswappeddict[cell] = onemicron_nodes
    return coordswappeddict
                
            

def get_target_nodes_list(nodes, *regions):
    '''
    Docstring for get_target_nodes_list
    
    :param nodes: Description
    :param regions: Description
    '''
    targets = []
    for node in nodes:
        if node['allenId'] in regions:
            targets.append(node)
    return targets

def get_coords(nodes, dim='all', mirror=False):
    """
    Coordinate getter for a list of nodes contianing x, y, z coordinates
    
    :param nodes: list of dictionaries, each entry should be a dictionary containing 'x', 'y', 'z' coordinates for said node
    :param dim: str, selecting which coordinates to get, default behavior is to return all coords
    :param mirror: bool, default False to not mirror nodes, but will mirror nodes over saggital axis (z in ccfv3)

    Returns: list of coordinates if only one dimension is selected, np.array of lists where each list contains x, y, z coords if all dims are selected
    """
    match dim:
        case 'x':
            x = [node['x'] for node in nodes]
            return x
        case 'y':
            y = [node['y'] for node in nodes]
            return y
        case 'z':
            if mirror:
                for node in nodes:
                    if node['z'] < 5700:
                        diff = 5700-node['z']
                        node['z'] = diff
                z = [node['z'] for node in nodes]
            else:
                z = [node['z'] for node in nodes]            
            return z
        case 'all':
            for node in nodes:
                if mirror:
                    if node['z'] < 5700:
                        diff = 5700 - node['z']
                        node['z'] = 5700+diff
            coords = np.array([[node['x'], node['y'], node['z']] for node in nodes]) 
            return coords
        
def node_coords_getter(cell, dim, mirror=False, *regions):
    targets = get_target_nodes_list(cell, regions)
    coords = np.array(get_coords(cell, dim, mirror) for cell in targets if cell)
    return coords

def get_cells_to_region(freqspkl, regionabv, thresh=3):
    freqs = pickle.load(open(freqspkl, 'rb'))

    latmerged = merge_regions(freqs)
    poscells = [cell for cell in latmerged.index.tolist() if latmerged.loc[cell][regionabv] > thresh]
    negcells = [cell for cell in latmerged.index.tolist() if latmerged.loc[cell][regionabv] <= thresh]

    return poscells, negcells

def get_targeted_regions(data, cell):
    '''
    returns a series containing the targeted regions of a given cell, requires data to be a pandas DataFrame where columns are regions and cells are rows
    cell is the name of the cell you want to see target regions for as str 
    '''
    return data.loc[cell, data.columns[data.loc[cell]!=0].tolist()].sort_values(ascending=False)

def swc_to_line_actors(swc_df, skip_dendrite=False, axon_color='blue', dendrite_color='red', soma_color='black', neurite_radius=4, soma_radius=15):

    """
    Build vedo Tube actors per morphological section directly from a
    parsed SWC DataFrame — no file path or morphio needed.

    Sections are extracted by walking parent-child relationships:
      - A section starts at any non-soma node whose parent is the soma,
        has no parent (parentNumber==-1), or is a branch point (2+ children)
      - The section walks forward along single-child nodes until hitting
        a tip (0 children) or branch point (2+ children)
      - The parent node coordinate is prepended to each child section so
        that adjacent tubes overlap at branch points, giving seamless joins
    """
    nodes = swc_df.set_index('id')

    # Build children map: parent_id -> [child_ids]
    children = defaultdict(list)
    for node_id, row in nodes.iterrows():
        if row['parent'] != -1:
            if skip_dendrite == False: 
                children[int(row['parent'])].append(node_id)
            else:
                if row['type'] == 3:
                    continue
                else:
                    children[int(row['parent'])].append(node_id)
    # Soma
    soma_row = nodes[nodes['type'] == 1].iloc[0]
    soma_id  = int(soma_row.name)
    soma_pos = soma_row[['x', 'y', 'z']].tolist()

    type_color_map = {2: axon_color, 3: dendrite_color}

    actors = [Sphere(pos=soma_pos, r=soma_radius, c=soma_color)]

    # Section start nodes: non-soma nodes whose parent is either
    # soma, -1 (orphan root), or a branch point (2+ children)
    section_starts = [
        node_id for node_id, row in nodes.iterrows()
        if row['type'] != 1
        and (   row['parent'] == -1
             or int(row['parent']) == soma_id
             or len(children[int(row['parent'])]) > 1)
    ]

    n_axon_sections = 0
    n_dend_sections = 0
    n_skipped       = 0

    for start_id in section_starts:
        section_ids = []

        # Prepend the branch-point parent coordinate so this tube overlaps
        # with the end of its parent tube — fills the gap at the branch point.
        # Skip if parent is soma (soma sphere already covers that join) or -1.
        parent_id = int(nodes.loc[start_id, 'parent'])
        if parent_id != -1 and parent_id != soma_id:
            section_ids.append(parent_id)

        # Walk the unbranched chain forward
        current_id = start_id
        while True:
            section_ids.append(current_id)
            kids = children[current_id]
            if len(kids) == 1:
                current_id = kids[0]
            else:
                break  # tip (0 children) or branch point (2+ children) — stop

        if len(section_ids) < 2:
            n_skipped += 1
            continue

        pts   = nodes.loc[section_ids, ['x', 'y', 'z']].values.tolist()
        ntype = int(nodes.loc[start_id, 'type'])
        color = type_color_map.get(ntype, axon_color)

        actors.append(Tube(pts, r=neurite_radius, c=color, cap=True))

        if ntype == 2:
            n_axon_sections += 1
        else:
            n_dend_sections += 1


    return actors
# =============================================================================
#     node_coords = swc_df.set_index('id')[['x', 'y', 'z']]
# 
#     # Only rows that have a valid parent
#     has_parent = swc_df[swc_df['parent'] != -1]
# 
#     actors = []
#     for neurite_type, color in [(2, axon_color), (3, dendrite_color)]:
#         subset = has_parent[has_parent['type'] == neurite_type]
#         if skip_dendrite and neurite_type == 3:
#             continue
#         if subset.empty:
#             continue
# 
#         # Build (N, 3) start and end point arrays
#         start_pts = node_coords.loc[subset['id']].values
#         end_pts   = node_coords.loc[subset['parent']].values
# 
#         actor = Lines(start_pts, end_pts, c=color, lw=lw)
#         actors.append(actor)
# 
#     return actors
# =============================================================================

def swap_for_brainrender(swcpath, axon='green', dendrite='black', soma='black', skip_dendrite=False, neurite_radius=4, soma_radius=15):
    """
    swap coordinates of an swc to be rendered with brainrender, when swc is acquired from the Allen Institute

    Parameters
    ----------
    swcpath : str
        path to the swc you need coordinates swapped for.

    Returns
    -------
    cell_actors : list
        a list of brainrender Line actors to be added to a Scene.

    """

    
    swc_df = pd.read_csv(
         swcpath,
         comment='#', 
         sep=r'\s+',
         names=['id', 'type', 'x', 'y', 'z', 'r', 'parent']
    )
    #need to swap x and z coordinates, brainrender swapped their axes so using swcs as we get them will render them rotated 90 deg
    x = swc_df['x']
    z = swc_df['z']
    swc_df['x'] = z
    swc_df['z'] = x
    cell_actors = swc_to_line_actors(
        swc_df, axon_color=axon, dendrite_color=dendrite, soma_color=soma, skip_dendrite=skip_dendrite, 
        neurite_radius=neurite_radius, soma_radius=soma_radius
        )
    return cell_actors

def write_targeted_regions_to_excel(data, output_path='targeted_regions.xlsx'):
    '''
    Runs get_targeted_regions for every cell in data and writes results to one Excel sheet.
    Each cell gets two adjacent columns: region names and their endpoint counts,
    under a merged header cell showing the cell name.

    Parameters
    ----------
    data        : pd.DataFrame  Regions as columns, cells as rows
    output_path : str           Path for the output .xlsx file
    '''

    # Build a plain dict of series for each cell
    col_data = {}
    for cell in data.index:
        series = get_targeted_regions(data, cell)
        col_data[cell] = {
            'regions':   series.index.tolist(),
            'endpoints': series.values.tolist()
        }

    wb = Workbook()
    ws = wb.active
    ws.title = 'Targeted Regions'

    cells = list(data.index)

    for i, cell in enumerate(cells):
        col_start = i * 2 + 1  # 1-indexed
        col_end   = col_start + 1

        # --- Row 1: merged cell name header ---
        ws.merge_cells(
            start_row=1, start_column=col_start,
            end_row=1,   end_column=col_end
        )
        header_cell = ws.cell(row=1, column=col_start, value=cell)
        header_cell.font      = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

        # --- Row 2: sub-headers ---
        region_header    = ws.cell(row=2, column=col_start, value='Region')
        endpoints_header = ws.cell(row=2, column=col_end,   value='Endpoints')
        region_header.font    = Font(bold=True)
        endpoints_header.font = Font(bold=True)

        # --- Row 3+: data ---
        regions   = col_data[cell]['regions']
        endpoints = col_data[cell]['endpoints']

        for row_offset, (region, endpoint) in enumerate(zip(regions, endpoints)):
            ws.cell(row=3 + row_offset, column=col_start, value=region)
            ws.cell(row=3 + row_offset, column=col_end,   value=endpoint)

    wb.save(output_path)
    print(f"Results written to '{output_path}'")


# --- Usage ---
# write_targeted_regions_to_excel(your_dataframe, output_path='targeted_regions.xlsx')